import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tempfile

st.set_page_config(page_title="ECOBSP", layout="centered")

# üîê Autenticaci√≥n b√°sica
st.title("üîí Acceso protegido - ECOBSP")
password = st.text_input("Introduce la contrase√±a:", type="password")
if password != "ecobsp2024":
    st.warning("Contrase√±a incorrecta. Acceso denegado.")
    st.stop()

st.markdown("## üìÑ Comparador BSP-ORBIS")
st.markdown("### Sube los archivos BSP (CSV) y ORBIS (Excel) para identificar billetes no encontrados en ORBIS.")

# üìå Subida de archivos
bsp_file = st.file_uploader("üîΩ Sube el archivo BSP (CSV)", type="csv")
orbis_file = st.file_uploader("üîΩ Sube el archivo ORBIS (Excel)", type=["xlsx", "xls"])

if bsp_file and orbis_file:
    try:
        # Leer BSP
        bsp_df = pd.read_csv(bsp_file, dtype=str, sep=';', header=0, encoding='utf-8')
        bsp_df.columns = bsp_df.columns.str.strip()

        # Leer ORBIS
        orbis_df = pd.read_excel(orbis_file, sheet_name=0, dtype=str, engine="openpyxl", header=0)
        orbis_df.columns = orbis_df.columns.str.strip()

        # Mostrar columnas para depuraci√≥n
        st.write("üìÑ Columnas del archivo BSP:")
        st.write(bsp_df.columns.tolist())
        st.write("üìä Columnas del archivo ORBIS:")
        st.write(orbis_df.columns.tolist())

        # Filtrar solo TKTT
        bsp_df = bsp_df[bsp_df["DOC 1A"] == "TKTT"]

        # Procesar n√∫mero de billete
        bsp_df["N_BILLETE_PROCESADO"] = bsp_df["N BILLETE"].str.strip().str[-10:]
        bsp_df["N_BILLETE_PROCESADO"] = pd.to_numeric(bsp_df["N_BILLETE_PROCESADO"], errors='coerce')

        # Extraer billetes de ORBIS
        orbis_numeros = pd.to_numeric(orbis_df["N¬∫ Billete"], errors='coerce').dropna().astype(int).tolist()
        bsp_df["ENCONTRADO"] = bsp_df["N_BILLETE_PROCESADO"].isin(orbis_numeros)

        # Mostrar resumen
        total_registros = len(bsp_df)
        registros_no_encontrados = len(bsp_df[~bsp_df["ENCONTRADO"]])

        st.success("‚úÖ Comparaci√≥n completada")
        st.write(f"üîç Total registros TKTT analizados: **{total_registros}**")
        st.write(f"‚ùå No encontrados en ORBIS: **{registros_no_encontrados}**")

        # Crear Excel de salida
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb = Workbook()
            ws = wb.active
            ws.append(bsp_df.columns.tolist())

            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            for _, row in bsp_df.iterrows():
                row_values = row.tolist()
                ws.append(row_values)
                if not row["ENCONTRADO"]:
                    for cell in ws[ws.max_row]:
                        cell.fill = red_fill
            wb.save(tmp.name)
            tmp_path = tmp.name

        with open(tmp_path, "rb") as f:
            st.download_button("üì• Descargar resultado en Excel", f, file_name="resultado_ECOBSP.xlsx")

    except Exception as e:
        st.error(f"‚ùå Error al procesar los archivos: {e}")

else:
    st.info("‚¨ÜÔ∏è Esperando que subas ambos archivos BSP y ORBIS.")

